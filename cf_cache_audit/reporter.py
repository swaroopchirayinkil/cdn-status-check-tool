"""Rich terminal reporter and JSON/CSV export for cf-cache-audit.

Renders:
  • Cloudflare detection banner
  • Asset table with colour-coded audit results
  • Warm-cache transition table (when ``--warm-cache`` is used)
  • Summary statistics panel
  • JSON and CSV exports
"""

from __future__ import annotations

import csv
import json
import logging
from datetime import datetime, timezone
from io import StringIO
from pathlib import Path
from typing import Any

from rich.box import ROUNDED
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.text import Text

from cf_cache_audit.models import (
    AssetInfo,
    AuditReport,
    AuditResult,
    AuditSummary,
    CfCacheStatus,
    CloudflareInfo,
)
from cf_cache_audit.utils import truncate

logger = logging.getLogger("cf_cache_audit")

# Rich console singleton
console = Console()

# Colour map for audit results
_RESULT_STYLES: dict[str, str] = {
    AuditResult.PASS.value: "bold green",
    AuditResult.WARNING.value: "bold yellow",
    AuditResult.ERROR.value: "bold red",
    AuditResult.INFO.value: "dim",
}

_CF_STATUS_STYLES: dict[str, str] = {
    CfCacheStatus.HIT.value: "green",
    CfCacheStatus.MISS.value: "yellow",
    CfCacheStatus.EXPIRED.value: "yellow",
    CfCacheStatus.REVALIDATED.value: "cyan",
    CfCacheStatus.STALE.value: "yellow",
    CfCacheStatus.BYPASS.value: "red",
    CfCacheStatus.DYNAMIC.value: "magenta",
    CfCacheStatus.UNKNOWN.value: "dim",
    CfCacheStatus.NONE.value: "dim",
}


# ---------------------------------------------------------------------------
# Cloudflare banner
# ---------------------------------------------------------------------------

def print_cloudflare_status(cf: CloudflareInfo) -> None:
    """Print a prominent Cloudflare detection banner."""
    if cf.detected:
        status = Text("✔ YES", style="bold green")
    else:
        status = Text("✘ NO", style="bold red")

    lines: list[str] = []
    if cf.server_header:
        lines.append(f"  Server header : {cf.server_header}")
    if cf.cf_ray_present:
        lines.append("  CF-Ray        : present")
    if cf.apo_detected:
        lines.append("  APO           : ✔ detected")
    for hint in cf.cache_rules_hints:
        lines.append(f"  Cache Rule    : {hint}")
    for note in cf.additional_notes:
        lines.append(f"  Note          : {note}")

    body = "\n".join(lines) if lines else ""

    panel = Panel(
        Text.assemble("Cloudflare detected: ", status, "\n", body),
        title="[bold cyan]☁  Cloudflare Detection[/bold cyan]",
        border_style="cyan",
        box=ROUNDED,
        expand=False,
        padding=(1, 2),
    )
    console.print(panel)
    console.print()


# ---------------------------------------------------------------------------
# Asset table
# ---------------------------------------------------------------------------

def print_asset_table(assets: list[AssetInfo], *, verbose: bool = False) -> None:
    """Render the main asset-status table."""
    table = Table(
        title="[bold]Asset Cache Status Report[/bold]",
        box=ROUNDED,
        show_lines=False,
        header_style="bold bright_white on grey23",
        row_styles=["", "on grey7"],
        expand=True,
        show_edge=True,
    )

    table.add_column("#", justify="right", style="dim", width=5)
    table.add_column("URL", style="cyan", ratio=4, no_wrap=True, overflow="fold")
    table.add_column("TYPE", style="white", width=12)
    table.add_column("STATUS", justify="center", width=7)
    table.add_column("CF CACHE", justify="center", width=14)
    table.add_column("CACHEABLE", justify="center", width=10)
    table.add_column("CDN", justify="center", width=12)
    table.add_column("RESULT", justify="center", width=10)

    if verbose:
        table.add_column("MESSAGE", ratio=2, no_wrap=False)

    # Sort: warnings first, then errors, then pass, then info
    _order = {
        AuditResult.WARNING.value: 0,
        AuditResult.ERROR.value: 1,
        AuditResult.PASS.value: 2,
        AuditResult.INFO.value: 3,
    }
    sorted_assets = sorted(assets, key=lambda a: _order.get(a.audit_result, 9))

    for idx, asset in enumerate(sorted_assets, 1):
        cf_style = _CF_STATUS_STYLES.get(asset.cf_cache_status, "dim")
        result_style = _RESULT_STYLES.get(asset.audit_result, "dim")

        cacheable_text = Text("YES", style="green") if asset.is_cacheable else Text("no", style="dim")

        http_status = str(asset.http_status) if asset.http_status else "—"

        row: list[Any] = [
            str(idx),
            asset.url,
            asset.asset_type,
            http_status,
            Text(asset.cf_cache_status, style=cf_style),
            cacheable_text,
            asset.cdn_provider or "—",
            Text(asset.audit_result, style=result_style),
        ]
        if verbose:
            row.append(asset.audit_message or "")

        table.add_row(*row)

    console.print(table)
    console.print()


# ---------------------------------------------------------------------------
# Warm-cache table
# ---------------------------------------------------------------------------

def print_warm_cache_table(assets: list[AssetInfo]) -> None:
    """Render warm-cache transition results for cacheable assets."""
    warm_assets = [a for a in assets if a.warm_cache_results]
    if not warm_assets:
        return

    table = Table(
        title="[bold]Warm-Cache Results[/bold]",
        box=ROUNDED,
        header_style="bold bright_white on grey23",
        expand=True,
    )
    table.add_column("URL", style="cyan", ratio=3, no_wrap=False)
    table.add_column("Attempts", ratio=3)
    table.add_column("Warmed?", justify="center", width=10)

    for asset in warm_assets:
        attempts_parts: list[str] = []
        statuses: list[str] = []
        for wa in asset.warm_cache_results:
            cf_val = wa.cf_cache_status if isinstance(wa.cf_cache_status, str) else wa.cf_cache_status.value
            attempts_parts.append(
                f"#{wa.attempt}: {cf_val} "
                f"({wa.response_time_ms:.0f}ms)"
            )
            statuses.append(cf_val)

        # Determine if cache warmed up (MISS → HIT)
        warmed = "—"
        if len(statuses) >= 2:
            if statuses[0] in ("MISS", "EXPIRED") and statuses[-1] == "HIT":
                warmed = "[green]✔ YES[/green]"
            elif statuses[-1] == "HIT":
                warmed = "[green]✔ YES[/green]"
            elif all(s == "HIT" for s in statuses):
                warmed = "[green]✔ Already cached[/green]"
            else:
                warmed = "[yellow]✘ NO[/yellow]"

        table.add_row(
            truncate(asset.url, 70),
            "\n".join(attempts_parts),
            warmed,
        )

    console.print(table)
    console.print()


# ---------------------------------------------------------------------------
# Summary panel
# ---------------------------------------------------------------------------

def print_summary(summary: AuditSummary) -> None:
    """Render the audit summary panel."""
    # Hit ratio colour
    ratio = summary.hit_ratio
    if ratio >= 80:
        ratio_style = "bold green"
    elif ratio >= 50:
        ratio_style = "bold yellow"
    else:
        ratio_style = "bold red"

    lines = [
        f"  Total assets      : [bold]{summary.total_assets}[/bold]",
        f"  Cacheable assets  : [bold]{summary.cacheable_assets}[/bold]",
        "",
        f"  [green]HIT[/green]             : {summary.hit}",
        f"  [yellow]MISS[/yellow]            : {summary.miss}",
        f"  [yellow]EXPIRED[/yellow]         : {summary.expired}",
        f"  [cyan]REVALIDATED[/cyan]     : {summary.revalidated}",
        f"  [yellow]STALE[/yellow]           : {summary.stale}",
        f"  [red]BYPASS[/red]          : {summary.bypass}",
        f"  [magenta]DYNAMIC[/magenta]         : {summary.dynamic}",
        f"  [dim]UNKNOWN[/dim]         : {summary.unknown}",
        f"  [dim]NO STATUS[/dim]       : {summary.none_status}",
        "",
        f"  Errors            : [red]{summary.errors}[/red]",
        f"  Warnings          : [yellow]{summary.warnings}[/yellow]",
        "",
        f"  Hit ratio         : [{ratio_style}]{ratio:.1f}%[/{ratio_style}]",
    ]

    fw = summary.framework_hint
    fw_str = fw.value if hasattr(fw, 'value') else str(fw)
    if fw_str and fw_str != "unknown":
        lines.append(f"  Framework         : {fw_str}")

    if summary.cdn_breakdown:
        lines.append("")
        lines.append("  CDN breakdown:")
        for cdn, count in sorted(
            summary.cdn_breakdown.items(), key=lambda x: -x[1]
        ):
            cdn_str = cdn.value if hasattr(cdn, 'value') else str(cdn)
            lines.append(f"    {cdn_str:<16}: {count}")

    panel = Panel(
        "\n".join(lines),
        title="[bold cyan]📊  Audit Summary[/bold cyan]",
        border_style="cyan",
        box=ROUNDED,
        expand=False,
        padding=(1, 2),
    )
    console.print(panel)
    console.print()


# ---------------------------------------------------------------------------
# JSON export
# ---------------------------------------------------------------------------

def export_json(report: AuditReport, path: str) -> None:
    """Write the full audit report to a JSON file."""
    data = report.model_dump(mode="json")
    Path(path).write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
    console.print(f"[green]✔[/green] JSON report saved to [bold]{path}[/bold]")


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

def export_csv(report: AuditReport, path: str) -> None:
    """Write asset-level data to a CSV file."""
    fieldnames = [
        "url",
        "asset_type",
        "http_status",
        "content_type",
        "content_length",
        "cache_control",
        "etag",
        "last_modified",
        "age",
        "cf_cache_status",
        "cf_ray",
        "cdn_provider",
        "is_cacheable",
        "audit_result",
        "audit_message",
        "error",
    ]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for asset in report.assets:
            row = {k: getattr(asset, k, "") for k in fieldnames}
            writer.writerow(row)

    console.print(f"[green]✔[/green] CSV report saved to [bold]{path}[/bold]")


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_xlsx(report: AuditReport, path: str) -> None:
    """Write asset-level data to an Excel workbook with formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Assets sheet ----------------------------------------------------
    ws = wb.active
    ws.title = "Assets"  # type: ignore[union-attr]

    headers = [
        "#",
        "URL",
        "Asset Type",
        "HTTP Status",
        "Content-Type",
        "Content-Length",
        "Cache-Control",
        "ETag",
        "Last-Modified",
        "Age",
        "CF-Cache-Status",
        "CF-Ray",
        "CDN Provider",
        "Cacheable",
        "Audit Result",
        "Audit Message",
        "Error",
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)  # type: ignore[union-attr]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Result colour fills
    fill_pass = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_warning = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_error = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    fill_info = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    result_fills = {
        "PASS": fill_pass,
        "WARNING": fill_warning,
        "ERROR": fill_error,
        "INFO": fill_info,
    }

    url_font = Font(color="0563C1", underline="single")

    for row_idx, asset in enumerate(report.assets, 2):
        row_data = [
            row_idx - 1,
            asset.url,
            asset.asset_type,
            asset.http_status,
            asset.content_type or "",
            asset.content_length,
            asset.cache_control or "",
            asset.etag or "",
            asset.last_modified or "",
            asset.age or "",
            asset.cf_cache_status,
            asset.cf_ray or "",
            asset.cdn_provider or "",
            "YES" if asset.is_cacheable else "NO",
            asset.audit_result,
            asset.audit_message or "",
            asset.error or "",
        ]

        row_fill = result_fills.get(asset.audit_result, fill_info)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)  # type: ignore[union-attr]
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center")

        # Make URL a clickable hyperlink
        url_cell = ws.cell(row=row_idx, column=2)  # type: ignore[union-attr]
        url_cell.font = url_font
        try:
            url_cell.hyperlink = asset.url
        except Exception:
            pass  # some URLs may not be valid hyperlinks for openpyxl

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 2:  # URL column — give generous width
            ws.column_dimensions[col_letter].width = 80  # type: ignore[union-attr]
        else:
            max_len = len(headers[col_idx - 1])
            for row_idx in range(2, min(len(report.assets) + 2, 102)):  # sample first 100 rows
                val = ws.cell(row=row_idx, column=col_idx).value  # type: ignore[union-attr]
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)  # type: ignore[union-attr]

    # Freeze top row
    ws.freeze_panes = "A2"  # type: ignore[union-attr]
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(report.assets) + 1}"  # type: ignore[union-attr]

    # ---- Summary sheet ---------------------------------------------------
    ws_summary = wb.create_sheet("Summary")
    summary = report.summary

    summary_data = [
        ("Website", report.website),
        ("Scan Started", report.scan_started),
        ("Scan Finished", report.scan_finished or ""),
        ("", ""),
        ("Cloudflare Detected", "YES" if report.cloudflare.detected else "NO"),
        ("Server Header", report.cloudflare.server_header or ""),
        ("CF-Ray Present", "YES" if report.cloudflare.cf_ray_present else "NO"),
        ("APO Detected", "YES" if report.cloudflare.apo_detected else "NO"),
        ("", ""),
        ("Total Assets", summary.total_assets),
        ("Cacheable Assets", summary.cacheable_assets),
        ("", ""),
        ("HIT", summary.hit),
        ("MISS", summary.miss),
        ("EXPIRED", summary.expired),
        ("REVALIDATED", summary.revalidated),
        ("STALE", summary.stale),
        ("BYPASS", summary.bypass),
        ("DYNAMIC", summary.dynamic),
        ("UNKNOWN", summary.unknown),
        ("NO STATUS", summary.none_status),
        ("", ""),
        ("Errors", summary.errors),
        ("Warnings", summary.warnings),
        ("Hit Ratio (%)", summary.hit_ratio),
    ]

    label_font = Font(bold=True, size=11)
    for row_idx, (label, value) in enumerate(summary_data, 1):
        label_cell = ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)
        if label:
            label_cell.font = label_font

    ws_summary.column_dimensions["A"].width = 22
    ws_summary.column_dimensions["B"].width = 50

    # ---- Save ------------------------------------------------------------
    wb.save(path)
    console.print(f"[green]✔[/green] Excel report saved to [bold]{path}[/bold]")
